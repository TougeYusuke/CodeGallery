import csv
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
import pandas as pd
import time
import openpyxl

# 排他的論理和
def xor(a,b):
    return bool(a) != bool(b)

# 無効な文字をアンダースコアに置き換える
def clean_sheet_name(name):
    invalid_chars = ['[', ']', ':', '*', '?', '/', '\\']
    cleaned_name = ''.join('_' if char in invalid_chars else char for char in name)
    return cleaned_name

# 対象のインデックスを取得
def find_to_index(list, value):
    if value in list:
        return value.index(value)
    else:
        return -1
    
# テキストの空判定
def is_empty_string(string):
    return not string
    
# サイトを開いた際の設定
def site_setting():
    # サーバーの設定
    combo_box_element = driver.find_element(By.ID, "servers") 
    select = Select(combo_box_element)
    select.select_by_value("Mandragora")

    #言語の設定
    combo_box_element = driver.find_element(By.ID, "languages") 
    select = Select(combo_box_element)
    select.select_by_value("ja")

    #タイムゾーンの設定
    combo_box_element = driver.find_element(By.ID, "timezones") 
    select = Select(combo_box_element)
    select.select_by_value("Asia/Tokyo")

    #左のリスト表示設定
    combo_box_element = driver.find_element(By.ID, "leftnav") 
    select = Select(combo_box_element)
    select.select_by_value("on")

    #保存ボタンをクリック
    save_btn_element = driver.find_element(By.CLASS_NAME, "btn-green") 
    save_btn_element.click()

    time.sleep(2)

    #自分のサーバーをクリック
    home_world_btn_element = driver.find_element(By.CLASS_NAME, "home-world") 
    home_world_btn_element.click()    

def get_category_list(url_category_text):
    # 各カテゴリテキストを取得
    category_btn_element_list = []
    category_btn_element_list = driver.find_elements(By.CLASS_NAME, url_category_text)

    return category_btn_element_list

# URLリストの取得
def get_url_list(category_btn_element, excel_worksheet_data):
    # 読み込みに失敗したら、ファイルの生成を行なう
    url_list = []

    # カテゴリボタンをクリックし、項目を表示
    category_btn_element.click()
    
    #表示された内容から、アイテムのURLを取得
    item_link_elements = driver.find_elements(By.TAG_NAME, 'a')
    for item_link in item_link_elements:
        href_value = item_link.get_attribute("href")

        # アイテムリンクでないものはスキップする
        searchText = 'market'
        if href_value and searchText in href_value:
            url_list.append(href_value)
        
    for url in url_list:
        # URLキャッシュデータに保存
        excel_worksheet_data.append([category_btn_element.text.strip(), url])
    
    return url_list

# アイテム名の取得
def get_item_name():
    # 要素が表示されるまで待つ
    wait = WebDriverWait(driver, 10)
    item_name_xpath = '/html/body/div/div/div[1]/main/div/div/div[1]/div[1]/div[3]/div[1]/h1'
    item_name_element = wait.until(EC.presence_of_element_located((By.XPATH, item_name_xpath)))
    return item_name_element.text.strip()

# エクセルファイルのリセット
def excel_data_reset():
    wb = openpyxl.Workbook()
    ws = wb.active

    # ファイル保存
    wb.save(excel_file_name)
    wb.close()

# 価格の取得
def gat_price(parent_element, is_hq):
    hq_mark_element = parent_element.find_element(By.CSS_SELECTOR, ".price-hq")
    if xor(is_hq, len(hq_mark_element.find_elements(By.TAG_NAME, "span"))>0):
        # アイテムが存在しない
        price = '-1'
    else:
        # 価格を取得
        price = elems.text.strip()

    return price

# 最後にアクセスしたアドレスを保存
def load_last_access_address():
    result = ""
    try:
        file = open(save_file_name, "r")
        result = file.readline()
        file.close()
    except:
        # 読み込みに失敗したら、ファイルの生成を行なう
        file = open(save_file_name, "w")
        file.close()

    return result

# 最後にアクセスしたアドレスを保存
def save_last_access_address(url):
    file = open(save_file_name, "w")
    file.write(url)
    file.close()

excel_file_name = "MarketBoudePriceList.xlsx"
save_file_name = "savedata.txt"
url_list_file_name = "url_list.xlsx"

# WebDriverのインスタンスを作成（Chromeを使用する例）
# ChromeDriverがインストールされている必要があります
driver = webdriver.Chrome()

# 適当なアイテムのURLを設定
start_url = load_last_access_address()

# 再開ではない場合は、エクセルデータを初期化する
if not start_url:
    excel_data_reset()

# とりあえず、適当なページを開く
driver.get("https://universalis.app/market/34563")

time.sleep(3)

# 初回アクセス時のサイトの初期設定
site_setting()

url_category_name_list = []
url_list = []

try:
    wb = openpyxl.load_workbook(url_list_file_name)
    ws = wb.active

    prev_category_name = ""
    load_url_list = []

    # エクセルデータを解析
    for row in ws.iter_rows(min_row=2, values_only=True):
        category  = row[0]
        url = row[1]

        # カテゴリが切り替わったら保存
        if category != prev_category_name:
            url_category_name_list.append(category)
            prev_category_name = category

            # 1つ以上項目が設定されている場合のみ対象
            if len(load_url_list)>0:
                # 読み込み終わっているURLをリストへ登録
                url_list.append(load_url_list)
                load_url_list.clear()

        load_url_list.append(url)

    # すべてのデータを読み込み終わったら、読み込み済みのurlをリストへ登録
    url_list.append(load_url_list)
    
    wb.close()

except:
    # アイテムのURLを取得
    url_class_list = ["type-weapons", "type-armor", "type-items", "type-housing"]

    # エクセルファイルにデータを出力
    try:
        wb = openpyxl.load_workbook(excel_file_name)
    except:
        wb = openpyxl.Workbook()        

    ws = wb.active
    ws.append(["Category", "URL"])

    for url_class in url_class_list:
        #カテゴリリストを取得
        category_element_list = get_category_list(url_class)

        for category_element in category_element_list:
            # URLリストの取得
            url_list.append(get_url_list(category_element, ws))
            # カテゴリ名を保存
            url_category_name_list.append(category_element.text.strip())
        
    # ファイル保存
    wb.save(url_list_file_name)
    wb.close()


category_index = 0
comeback = is_empty_string(start_url)

# 数が多いのでカテゴリごとに分けて出力
for item_category_url_list in url_list:
    data_list = []

    for url in item_category_url_list:
        # 指定のファイルから再開
        if not comeback and url != start_url:
            continue
        comeback = True
        
        # ページを開く
        driver.get(url)

        # アイテム名を取得
        item_name = get_item_name()

        #カテゴリの取得
        category_name_elem = driver.find_element(By.CLASS_NAME, 'item_info2')
        category_name = category_name_elem.text.split(); 

        wait = WebDriverWait(driver, 10)

        try:
            # 価格一覧を取得
            class_name = '.price-current'                   # class属性名
            prive_list_elem = driver.find_element(By.XPATH, '/html/body/div[1]/div/div[1]/main/div/div/div[2]/div[7]/div[1]/div[1]/div/table')
            class_elems = prive_list_elem.find_elements(By.CSS_SELECTOR, class_name) 

            # 出品がない
            if len(class_elems) <= 0:
                hq_price = -1
                nq_price = -1
            else:
                #HQソートサーバーをクリック    
                hq_sort_xpath = '/html/body/div[1]/div/div[1]/main/div/div/div[2]/div[7]/div[1]/div[1]/div/table/thead/tr/th[2]'
                hq_sort_element = wait.until(EC.presence_of_element_located((By.XPATH, hq_sort_xpath)))
                hq_sort_element.click()
                time.sleep(1)

                # HQアイテムを取得
                for elems in class_elems:
                    # テキストが空なら何もしない
                    if elems.text.strip():
                        parent_element = elems.find_element(By.XPATH, "..")
                        # 価格を取得
                        hq_price = gat_price(parent_element, True)
                        # 1つ取得したら終了
                        break

                # HQソートサーバーをクリック
                hq_sort_element.click()
                time.sleep(1)

                # NQアイテムを取得
                for elems in class_elems:
                    # テキストが空なら何もしない
                    if elems.text.strip():
                        parent_element = elems.find_element(By.XPATH, "..")
                        # 価格を取得
                        nq_price = gat_price(parent_element, False)
                        # 1つ取得したら終了
                        break
            
            prive_list_elem = driver.find_element(By.XPATH, '/html/body/div[1]/div/div[1]/main/div/div/div[2]/div[7]/div[1]/div[2]/div/table')
            class_elems = prive_list_elem.find_elements(By.CSS_SELECTOR, class_name) 

            # 出品がない
            if len(class_elems) <= 0:
                histry_hq_price = -1
                histry_nq_price = -1
            else:
                #HQソートサーバーをクリック    
                history_hq_sort_xpath = '/html/body/div[1]/div/div[1]/main/div/div/div[2]/div[7]/div[1]/div[2]/div/table/thead/tr/th[2]'
                history_hq_sort_element = wait.until(EC.presence_of_element_located((By.XPATH, history_hq_sort_xpath)))
                history_hq_sort_element.click()
                time.sleep(1)

                # HQアイテムを取得
                for elems in class_elems:
                    # テキストが空なら何もしない
                    if elems.text.strip():
                        parent_element = elems.find_element(By.XPATH, "..")
                        # 価格を取得
                        histry_hq_price = gat_price(parent_element, True)

                        #最終取引時間の取得
                        histry_date_element = parent_element.find_element(By.CSS_SELECTOR, ".price-date")
                        histry_date_hq = histry_date_element.text.strip()
                        
                        #1つ取得したら終了
                        break

                #HQソートサーバーをクリック
                history_hq_sort_element.click()
                time.sleep(1)

                # NQアイテムを取得
                for elems in class_elems:
                        # テキストが空なら何もしない
                    if elems.text.strip():
                        parent_element = elems.find_element(By.XPATH, "..")
                        # 価格を取得
                        histry_nq_price = gat_price(parent_element, True)

                        #最終取引時間の取得
                        histry_date_element = parent_element.find_element(By.CSS_SELECTOR, ".price-date")
                        histry_date_nq = histry_date_element.text.strip()

                        #1つ取得したら終了
                        break
        except:
            hq_price = -1
            nq_price = -1
            histry_hq_price = -1
            histry_nq_price = -1
            histry_date_hq = ""

        # リストデータに登録
        data_list.append((item_name + '_HQ', category_name[0], category_name[2], hq_price, histry_hq_price, histry_date_hq))
        data_list.append((item_name + '_NQ', category_name[0], category_name[2], nq_price, histry_nq_price, histry_date_nq))

        save_last_access_address(url)

    if len(data_list) > 0:
        # エクセルファイルにデータを出力
        wb = openpyxl.load_workbook(excel_file_name)

        # 不正な文字列を「_」へ置き換え
        sheet_name = clean_sheet_name(url_category_name_list[category_index])

        # シート名を取得
        sheet_name_list = wb.sheetnames
        index = find_to_index(sheet_name_list, sheet_name)

        if index < 0:
            # 新規シートを作成
            ws = wb.create_sheet(sheet_name)
        else:
            # シートを移動
            ws = wb.worksheets[index]

        ws.append(["アイテム名", "カテゴリ", "サブカテゴリ", "値段", "取引値段", "最終取引日時"])

        for item_data in data_list:
            ws.append(item_data)

        # ファイル保存
        wb.save(excel_file_name)
        wb.close()

    category_index += 1

# WebDriverを閉じる
driver.quit()